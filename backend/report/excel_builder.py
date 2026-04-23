"""
Excel report builder matching the target format:

Sheet 1 "比价矩阵":
  Rows : TYPE (DIRECT/HD/H) × Route × Cabin (Y then C)
  Cols : [TYPE, OD, CABIN] + per-date × [CA, CZ, MU, BA/VS, HU, lowest(other)]
  Date headers : "MMM YYYY" / "dMMM-dMMM" two-line groups

Sheet 2 "详细数据":
  航线 | 日期 | TYPE | 舱位 | CA | CZ | MU | BA/VS | HU |
  市场最低价 | 最低价航司 | lowest fare(other) | 其他最低航司 | 国航指数 | 国航排名
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from report.analytics import (
    CA_CODE, build_excel_data, build_period_excel_data, enrich_flights,
    format_date_period, summary_stats,
)

# ── Palette ───────────────────────────────────────────────────────────────────
C_NAVY       = "FF1A2E3F"
C_MID_BLUE   = "FF2C4A6C"
C_STEEL      = "FF3A6186"
C_DATE_HDR   = "FF4472C4"   # blue column-group header
C_CA_COL     = "FFFFF0F0"   # light red  — CA column
C_MIN_CELL   = "FFFFFF00"   # yellow     — row minimum
C_DIRECT_BG  = "FFF0F8FF"   # alice blue — DIRECT section rows
C_HD_BG      = "FFFFF8F0"   # linen      — HD section rows
C_H_BG       = "FFF0FFF0"   # honeydew   — H section rows
C_WHITE      = "FFFFFFFF"
C_BLACK      = "FF000000"
C_GRAY       = "FFAAAAAA"
C_BORDER     = "FFCCCCCC"

TYPE_BG = {"DIRECT": "FF1F4E79", "ID": "FF375623", "II": "FF7B2C2C"}
ROW_BG  = {"DIRECT": C_DIRECT_BG, "ID": C_HD_BG, "II": C_H_BG}


def _fill(h: str) -> PatternFill:
    return PatternFill(start_color=h, end_color=h, fill_type="solid")

def _font(color=C_BLACK, bold=False, size=10):
    return Font(color=color, bold=bold, size=size, name="Arial")

def _border():
    t = Side(style="thin", color=C_BORDER)
    return Border(left=t, right=t, top=t, bottom=t)

def _c(align="center"):
    return Alignment(horizontal=align, vertical="center", wrap_text=False)

def _set(cell, value=None, font=None, fill=None, align="center"):
    if value is not None:
        cell.value = value
    cell.font   = font  or _font()
    cell.fill   = fill  or _fill(C_WHITE)
    cell.alignment = _c(align)
    cell.border = _border()


# ── Sheet 1: 比价矩阵 ─────────────────────────────────────────────────────────

_COL_FIXED = 3          # TYPE, OD, CABIN
_COLS_PER_DATE = 6      # CA, CZ, MU, BA/VS, HU, lowest(other)
_SUB_HDRS = ["CA", "CZ", "MU", "BA/VS", "HU", "lowest fare(other)"]


def _matrix_col(date_idx: int, sub_idx: int) -> int:
    """1-based column for a given date-group and sub-column index (0-5)."""
    return _COL_FIXED + 1 + date_idx * _COLS_PER_DATE + sub_idx


def _write_matrix_sheet(ws, data: dict, currency: str) -> None:
    dates      = data["dates"]
    matrix_rows = data["matrix_rows"]
    n_dates    = len(dates)
    total_cols = _COL_FIXED + n_dates * _COLS_PER_DATE

    ws.title = "比价矩阵"

    # ── Row 1: fixed headers + date-group merged headers ──────────────────────
    for col, lbl in [(1, "TYPE"), (2, "OD (CITY PAIR)"), (3, "CABIN")]:
        c = ws.cell(1, col)
        _set(c, lbl, font=_font(C_WHITE, bold=True, size=11), fill=_fill(C_NAVY))
        ws.cell(2, col).border = _border()
        ws.cell(2, col).fill   = _fill(C_NAVY)

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)

    date_labels = data.get("date_labels", {})   # period mode: pre-computed labels

    for di, date in enumerate(dates):
        if date in date_labels:
            month_yr, week_rng = date_labels[date]
        else:
            month_yr, week_rng = format_date_period(date, all_dates=dates)
        col_start = _matrix_col(di, 0)
        col_end   = _matrix_col(di, _COLS_PER_DATE - 1)

        # Row 1: month+year span
        c1 = ws.cell(1, col_start)
        _set(c1, f"{month_yr}\n{week_rng}",
             font=_font(C_WHITE, bold=True, size=10),
             fill=_fill(C_DATE_HDR))
        c1.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        ws.merge_cells(start_row=1, start_column=col_start,
                       end_row=1,   end_column=col_end)

        # Row 2: sub-headers
        for si, sh in enumerate(_SUB_HDRS):
            c2 = ws.cell(2, col_start + si)
            _set(c2, sh, font=_font(C_WHITE, bold=True, size=9),
                 fill=_fill(C_MID_BLUE))

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 7
    for di in range(n_dates):
        for si in range(_COLS_PER_DATE):
            letter = get_column_letter(_matrix_col(di, si))
            ws.column_dimensions[letter].width = 10 if si < 5 else 16

    ws.freeze_panes = "D3"

    # ── Data rows ─────────────────────────────────────────────────────────────
    current_type   = None
    type_start_row = None
    prev_route     = None
    route_start_row = None
    data_row       = 3

    def _close_type_merge(end_row):
        if type_start_row and end_row > type_start_row:
            ws.merge_cells(start_row=type_start_row, start_column=1,
                           end_row=end_row - 1,       end_column=1)

    def _close_route_merge(end_row):
        if route_start_row and end_row > route_start_row:
            ws.merge_cells(start_row=route_start_row, start_column=2,
                           end_row=end_row - 1,        end_column=2)

    for mrow in matrix_rows:
        type_lbl  = mrow["type_label"]
        route     = mrow["route"]
        cabin     = mrow["cabin"]
        date_data = mrow["date_data"]
        row_fill  = _fill(ROW_BG.get(type_lbl, C_WHITE))

        # Type change
        if type_lbl != current_type:
            if current_type is not None:
                _close_type_merge(data_row)
                _close_route_merge(data_row)
            current_type    = type_lbl
            type_start_row  = data_row
            prev_route      = None

        # Route change
        if route != prev_route:
            if prev_route is not None:
                _close_route_merge(data_row)
            prev_route      = route
            route_start_row = data_row

        # Col A: TYPE (will be merged later)
        c = ws.cell(data_row, 1)
        _set(c, type_lbl,
             font=_font(C_WHITE, bold=True, size=10),
             fill=_fill(TYPE_BG.get(type_lbl, C_NAVY)))

        # Col B: OD
        c = ws.cell(data_row, 2)
        _set(c, route, font=_font(bold=True), fill=row_fill)

        # Col C: CABIN
        c = ws.cell(data_row, 3)
        _set(c, cabin, font=_font(), fill=row_fill)

        # Date columns
        for di, date in enumerate(dates):
            dd = date_data.get(date)

            # Collect raw prices for this row+date to find minimum
            row_prices: list[float] = []
            if dd:
                for k in ("CA", "CZ", "MU", "HU"):
                    if dd.get(k): row_prices.append(dd[k])
                if dd.get("BA_VS"): row_prices.append(dd["BA_VS"][0])
                if dd.get("other"): row_prices.append(dd["other"][0])

            row_min = min(row_prices) if row_prices else None

            def _price_cell(col, price, suffix="", is_ca=False):
                c = ws.cell(data_row, col)
                if price:
                    is_min = row_min is not None and abs(price - row_min) < 0.01
                    val    = f"{round(price):,}"
                    if suffix:
                        val += f"-{suffix}"
                    if is_min:
                        _set(c, val, font=_font(bold=True), fill=_fill(C_MIN_CELL))
                    elif is_ca:
                        _set(c, val, font=_font(bold=True), fill=_fill(C_CA_COL))
                    else:
                        _set(c, val, font=_font(), fill=row_fill)
                else:
                    _set(c, "", font=_font(C_GRAY), fill=row_fill)

            base = _matrix_col(di, 0)
            if dd:
                _price_cell(base,     dd.get("CA"),  is_ca=True)
                _price_cell(base + 1, dd.get("CZ"))
                _price_cell(base + 2, dd.get("MU"))
                bv = dd.get("BA_VS")
                _price_cell(base + 3, bv[0] if bv else None, bv[1] if bv else "")
                _price_cell(base + 4, dd.get("HU"))
                ot = dd.get("other")
                c  = ws.cell(data_row, base + 5)
                if ot:
                    is_min = row_min is not None and abs(ot[0] - row_min) < 0.01
                    _set(c, f"{round(ot[0]):,}-{ot[1]}",
                         font=_font(bold=is_min),
                         fill=_fill(C_MIN_CELL) if is_min else row_fill)
                else:
                    _set(c, "", font=_font(C_GRAY), fill=row_fill)
            else:
                for si in range(_COLS_PER_DATE):
                    c = ws.cell(data_row, base + si)
                    _set(c, "", font=_font(C_GRAY), fill=row_fill)

        ws.row_dimensions[data_row].height = 15
        data_row += 1

    # Close last merges
    _close_type_merge(data_row)
    _close_route_merge(data_row)

    # Legend
    mode_note = ""
    if date_labels:
        mode_note = "（按时间段对比）"
    elif dates:
        from report.analytics import format_date_period as _fdp
        _, sample_label = _fdp(dates[0], all_dates=dates)
        if len(sample_label) <= 6 and "-" not in sample_label:
            mode_note = "（按日期对比）"
        elif "-" in sample_label:
            mode_note = "（按周对比）"
        else:
            mode_note = "（按月对比）"

    ws.cell(data_row + 1, 1).value = (
        f"价格单位：{currency}{mode_note}  |  黄色=当日最低价  红底=国航价格  "
        "lowest fare(other) = 全市场最低（含所有航司）  "
        "TYPE: DIRECT=直达  ID=国际出发→国内目的地  II=国际出发→国际目的地"
    )
    ws.cell(data_row + 1, 1).font = _font(C_GRAY, size=9)
    try:
        ws.merge_cells(start_row=data_row + 1, start_column=1,
                       end_row=data_row + 1, end_column=min(12, total_cols))
    except Exception:
        pass


# ── Sheet 2: 详细数据 ─────────────────────────────────────────────────────────

_DETAIL_HDRS = [
    ("航线",                    16),
    ("日期",                    13),
    ("TYPE",                     9),
    ("舱位",                     8),
    ("中国国航 CA",              11),
    ("南方航空 CZ",              11),
    ("东方航空 MU",              11),
    ("英航/维珍 BA/VS",         13),
    ("海南航空 HU",              11),
    ("lowest fare(other)",      18),   # global market minimum (any airline)
    ("最低价航司",               14),
    ("国航指数 (CA/min×100)",   18),
    ("国航排名",                  9),
]


def _write_detail_sheet(ws, detail_rows: list[dict], currency: str) -> None:
    ws.title = "详细数据"

    # Header
    for ci, (lbl, width) in enumerate(_DETAIL_HDRS, start=1):
        c = ws.cell(1, ci)
        _set(c, lbl, font=_font(C_WHITE, bold=True, size=10), fill=_fill(C_NAVY))
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_DETAIL_HDRS))}1"

    TYPE_BG_DETAIL = {"DIRECT": C_DIRECT_BG, "ID": C_HD_BG, "II": C_H_BG}

    for ri, row in enumerate(detail_rows, start=2):
        bg = _fill(TYPE_BG_DETAIL.get(row["type"], C_WHITE))

        def _val(v, fmt=None):
            if v is None:
                return ""
            if fmt == "price":
                return round(v, 0)
            return v

        market_min = row.get("market_min")
        ca_p = row.get("CA")

        # Highlight lowest-fare cell yellow if CA is the market minimum
        ca_is_min = (ca_p is not None and market_min is not None
                     and abs(ca_p - market_min) < 0.01)

        vals = [
            row["route"],
            row["date"],
            row["type"],
            row["cabin"],
            _val(ca_p, "price"),
            _val(row.get("CZ"), "price"),
            _val(row.get("MU"), "price"),
            (f"{round(row['BA_VS']):,}-{row['BA_VS_name']}"
             if row.get("BA_VS") else ""),
            _val(row.get("HU"), "price"),
            # lowest fare(other) = global market min, show price-airline
            (f"{round(market_min):,}-{row.get('min_airline','')}"
             if market_min else ""),
            row.get("min_airline", ""),
            _val(row.get("ca_index")),
            row.get("ca_all_rank", ""),
        ]

        for ci, val in enumerate(vals, start=1):
            c = ws.cell(ri, ci)
            c.value     = val if val != "" else None
            c.border    = _border()
            c.alignment = _c("center")

            # Col 5 = CA price
            if ci == 5 and ca_p is not None:
                c.font = _font(bold=True)
                c.fill = _fill(C_MIN_CELL) if ca_is_min else _fill(C_CA_COL)
            # Col 10 = lowest fare(other) = market min → always yellow
            elif ci == 10 and market_min:
                c.font = _font(bold=True)
                c.fill = _fill(C_MIN_CELL)
            else:
                c.font = _font()
                c.fill = bg

        ws.row_dimensions[ri].height = 15


# ── Public builder ────────────────────────────────────────────────────────────

def build_excel(
    output_path: Path,
    flights: list[dict],
    matrix: dict[str, Any],      # legacy arg — ignored
    stats: dict[str, Any],
    title: str,
    date_ranges: list[dict] | None = None,
) -> None:
    currency = stats.get("currency", "HKD")

    # Enrich flights (adds category, type_display)
    enriched = enrich_flights(flights)

    # Daily data — always used for 详细数据 sheet (raw per-day records)
    daily_data = build_excel_data(enriched)

    # Matrix data — use period aggregation when date_ranges provided, else daily
    if date_ranges:
        matrix_data = build_period_excel_data(enriched, date_ranges)
    else:
        matrix_data = daily_data

    wb = Workbook()
    ws_default = wb.active
    wb.remove(ws_default)

    # Sheet 1: 比价矩阵 (period-aggregated when ranges present, else daily)
    ws1 = wb.create_sheet("比价矩阵")
    _write_matrix_sheet(ws1, matrix_data, currency)

    # Sheet 2: 详细数据 (always raw per-day data)
    ws2 = wb.create_sheet("详细数据")
    _write_detail_sheet(ws2, daily_data["detail_rows"], currency)

    # Sheet 3: 报告信息
    ws3 = wb.create_sheet("报告信息")
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 42
    for r, (k, v) in enumerate([
        ("报告标题",   title),
        ("生成时间",   stats.get("generated_at", "")),
        ("日期范围",   stats.get("date_range", "")),
        ("币种",       currency),
        ("总航班数",   stats.get("total_flights", "")),
        ("航司数量",   stats.get("total_airlines", "")),
        ("最低价格",   stats.get("price_min", "")),
        ("平均价格",   stats.get("price_avg", "")),
    ], start=1):
        ws3.cell(r, 1, k).font = _font(bold=True)
        ws3.cell(r, 2, str(v))

    wb.save(output_path)
